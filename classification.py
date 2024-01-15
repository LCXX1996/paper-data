
#coding：utf-8
import sklearn
import nltk
import json
import math
import numpy as np
from math import log
from sklearn import metrics
from sklearn import svm, datasets
from sklearn.datasets import load_files
from sklearn.metrics import classification_report
from sklearn.feature_extraction.text import CountVectorizer, TfidfTransformer
from sklearn.feature_selection import SelectKBest,chi2
from sklearn.naive_bayes import MultinomialNB, BernoulliNB, GaussianNB
from openpyxl import Workbook
from openpyxl import load_workbook#引入库
from sklearn.preprocessing import Normalizer

print("=====读取训练集文本合并=====")
with open('训练集文本合并.txt') as file_object:
    patent_train = file_object.read()

print("=====读取测试集文本合并=====")
with open('测试集文本合并.txt') as file_object:
    patent_test = file_object.read()


print("=====建立数据集类别名列表=====")
patent_target_names=['稳定','产生','分离','吸附','调控','检测','结合','积累','引进','输出','移动']

print("=====训练样本分句=====")#参考文档《利用NLTK在Python下进行自然语言处理》,每个句子就是一篇文档
with open('训练集文本合并.txt') as file_object:
    paragraph = file_object.read()
# print(paragraph)
# print(type(paragraph))
sent_tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')
sentences = sent_tokenizer.tokenize(paragraph)
print(len(sentences))#9
print(type(sentences))#<class 'list'>


print("=====计算训练样本原始特征词的tfidf值，通过Scikit-Learn计算得到train_counts=====")
##依次读取停用词文本stop_words_list的每一行，将每一行作为一个停用词。
stop_list=[]#建立空列表用于存储停用词，以方便后面调用。
f = open("stop_words_list.txt", "r")
while True:
    stop_list_line = f.readline()
    if stop_list_line:
        pass    # do something here
        stop_list_line=stop_list_line.strip()
        p=stop_list_line.rfind('.')
        filename=stop_list_line[0:p]
        #print("create %s"%stop_list_line)
        stop_list.append(stop_list_line)
    else:
        break
f.close()
#print("停用词列表\n",stop_list)
print(len(stop_list))#1169
print(type(stop_list))#<class 'list'>


#计算词频
count_vect = CountVectorizer(stop_words=stop_list)
train_counts = count_vect.fit_transform(sentences).toarray()
print(train_counts.shape)#(9,280)


#以下程序段是提取训练集和测试集中各文本所属的类别名

print("--------以下是提取训练集中各文本所属的类别名--------")
train_target = []
for i in range(1,126):
    train_target.append('稳定')
for i in range(1,32):
    train_target.append('产生')
for i in range(1,120):
    train_target.append('分离')
for i in range(1,15):
    train_target.append('吸附')
for i in range(1,141):
    train_target.append('调控')
for i in range(1,44):
    train_target.append('检测')
for i in range(1,48):
    train_target.append('结合')
for i in range(1,29):
    train_target.append('积累')
for i in range(1,30):
    train_target.append('引进')
for i in range(1,19):
    train_target.append('输出')
for i in range(1,63):
    train_target.append('移动')

print("--------以下是提取测试集中各文本所属的类别名--------")
test_target = []
for i in range(1,32):
    test_target.append('稳定')
for i in range(1,9):
    test_target.append('产生')
for i in range(1,31):
    test_target.append('分离')
for i in range(1,4):
    test_target.append('吸附')
for i in range(1,35):
    test_target.append('调控')
for i in range(1,12):
    test_target.append('检测')
for i in range(1,13):
    test_target.append('结合')
for i in range(1,8):
    test_target.append('积累')
for i in range(1,8):
    test_target.append('引进')
for i in range(1,5):
    test_target.append('输出')
for i in range(1,16):
    test_target.append('移动')


print("=====保存训练样本原始特征词的字典=====")
filename = 'patent-train-stemmed_feature_names_dic.txt'#保存训练样本的X_train_counts
with open(filename, 'w',encoding='utf-8') as f_obj:#不能将'w'写成'wb'，否则会出错(wb是二进制读写)
    json.dump(str(count_vect.vocabulary_.items()), f_obj,ensure_ascii=False)#注意要转换成字符串str,不能转换成list

print("=====只保存训练样本原始特征词=====")
feature_names=count_vect.get_feature_names()
filename = 'patent-train-stemmed_feature_names.txt'#保存训练样本的X_train_counts
with open(filename, 'w',encoding='utf-8') as f_obj:#不能将'w'写成'wb'，否则会出错
    json.dump(feature_names, f_obj,ensure_ascii=False)
    

print("=====找到词频大于2的特征词的索引=====")
sum_columns=train_counts.sum(axis=0)#是计算矩阵每一列元素相加之和。
#print(sum_columns)
indexs=[]#空列表用于存储词频大于2的特征词的索引
i=0#索引从0开始
for sum_column, feature_name in zip(sum_columns,feature_names):#也可以不用for循环，用np.where(sum_columns>=2)，但是用该语句时，结果的类型为<class 'tuple'>
    if sum_column>=2 and len(feature_name)>1:
        indexs.append(i)
    i=i+1
print(type(indexs))#<class 'list'>
print(len(indexs))#96
#print(indexs)

print("=====根据索引，用一条语句去除词频小于2的特征词=====")
#count_vect_indexs={key:value for key,value in count_vect.vocabulary_.items() if value in indexs}# 结果是一个字典，该语句保持了特征词（键）和其位置信息（值），该方法大概用时20分钟。
#print(count_vect_indexs)

print("=====测试文本分句=====")#参考文档《利用NLTK在Python下进行自然语言处理》,每个句子就是一篇文档
#读取替换后的文件，相当于一个“段落”
with open('测试集文本合并.txt') as file_object:
    paragraph = file_object.read()
# print(paragraph)
# print(type(paragraph))
sent_tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')
sentences = sent_tokenizer.tokenize(paragraph)
print(type(sentences))#<class 'list'>


print("=====计算测试样本原始特征词的tfidf值，通过Scikit-Learn计算得到test_counts=====")
test_counts = count_vect.transform(sentences).toarray()
print(test_counts.shape)#(2189, 14575)



'''
以下两段程序是根据索引，先对train_counts（或test_counts ）切片以选择词频大于2的特征词，
再根据切片后的train_counts_indexs（test_counts_indexs）重新计算训练样本的TF-IDF
并且两段程序应该放在一起，因为计算测试样本的TF-IDF时要用到训练样本的fit方法
'''

print("=====根据索引，先对train_counts切片以选择词频大于2的特征词，再根据切片后的train_counts_indexs重新计算训练样本的TF-IDF=====")
train_counts_indexs=train_counts[:,indexs]
print(train_counts_indexs.shape)#(5485, 8786)

print("------将train_counts_indexs中大于1的元素用1替换，方便后面计算某特征词在多少篇文本中出现过------")
train_counts_indexs_replace=np.where(train_counts_indexs <= 1, train_counts_indexs, 1)#保留矩阵中小于等于1的元素不变，而其他元素都用1替换。参考《python numpy 数组中元素大于等于0的元素》20181010
#print(train_counts_indexs_replace)
print("-----对替换后的train_counts_indexs_replace按列求和，得到一个行向量，其中的每一个元素表示某特征词在多少篇文本中出现过-----")
train_counts_indexs_replace_sum=train_counts_indexs_replace.sum(axis=0)#是计算矩阵每一列元素相加之和。
#print(train_counts_indexs_replace_sum)
#print(type(train_counts_indexs_replace_sum))#<class 'numpy.ndarray'>

[train_rows,train_cols]=train_counts_indexs.shape#train_rows表示行数即文档的总数，cols表示列数即特征词的个数
N=train_rows#train_rows表示行数即文档的总数N
print(train_rows,train_cols)

print("=====根据索引，先对test_counts切片以选择词频大于2的特征词，再根据切片后的test_counts_indexs重新计算测试样本的TF-IDF=====")
test_counts_indexs=test_counts[:,indexs]
print(test_counts_indexs.shape)#(2189, 8786)

[test_rows,test_cols]=test_counts_indexs.shape#test_rows表示行数即文档的总数，cols表示列数即特征词的个数
#print(test_rows,test_cols)



'''建立用于存储特征（训练集和测试集）权重的矩阵，维数分别与train_counts_indexs和test_counts_indexs相同'''
train_tf_idf_indexs=np.zeros((train_rows,train_cols))#建立一个维数和train_counts相同的，且元素全为0的矩阵，用于存储计算的tf_idf值。也可以用print(train_counts.copy())建立矩阵
test_tf_idf_indexs=np.zeros((test_rows,test_cols))#建立一个维数和test_counts相同的，且元素全为0的矩阵，用于存储计算的测试集tf_idf值。注意用的是双括号。也可以用print(test_counts.copy())建立矩阵

print("=====计算训练集的特征权重=====")
for i in range(train_rows):
    for j in range(train_cols):
        if train_counts_indexs[i, j] != 0:
            train_tf_idf_indexs[i,j]=train_counts_indexs[i,j]*math.log(train_rows/train_counts_indexs_replace_sum[j])#math.log(c,b) #计算以b为底，c的对数:当参数b缺省，默认取自然对数(即底数，默认为 e).
    print("当前正在计算训练集的第",i+1,"个文本")

print("=====计算测试集的特征权重=====")
for i in range(test_rows):
    for j in range(test_cols):
        if test_counts_indexs[i, j] != 0:
            test_tf_idf_indexs[i,j]=test_counts_indexs[i,j]*math.log(train_rows/train_counts_indexs_replace_sum[j])#math.log(c,b) #计算以b为底，c的对数:当参数b缺省，默认取自然对数(即底数，默认为 e).
    print("当前正在计算测试集的第",i+1,"个文本")


print("-----对train_tf_idf_indexs归一化-----")


nm=Normalizer()
nm.fit(train_tf_idf_indexs)
train_tfidf_counts_indexs=nm.transform(train_tf_idf_indexs)#等价于上面两句
#print(train_tfidf_counts_indexs)
#print(train_tfidf_counts_indexs.shape)

print("-----对test_tf_idf_indexs归一化-----")
nm=Normalizer()
nm.fit(test_tf_idf_indexs)
test_tfidf_counts_indexs=nm.transform(test_tf_idf_indexs)
#print(test_tfidf_counts_indexs)
#print(type(test_tfidf_counts_indexs))#<class 'numpy.ndarray'>


'''释放Python的内存'''
import gc #garbage collector
del train_tf_idf_indexs,test_tf_idf_indexs,train_counts,test_counts,sum_columns,sentences,paragraph,patent_test,patent_train,stop_list,feature_names,train_counts_indexs_replace_sum
gc.collect()

'''--------------（分类器选择NB）用for循环从卡方检验中取不同的k个特征，即选出最优的k个特征用NB进行文本分类----------------'''
fs_num_list1 = list(range(100, 1100, 100))#在特征范围10-100内，间隔10取特征，得到10-90的特征数。Python3中range() 函数返回的是一个可迭代对象（类型是对象），而不是列表类型， 所以打印的时候不会打印列表。因此要加上list
fs_num_list2 = list(range(1000, 4300, 500))#取所有特征
fs_num_list3 = [4200]#取所有特征
fs_num_list = fs_num_list1 +fs_num_list2+fs_num_list3 #合并特征列表


f1_macro_average = list()#用于存储宏平均
f1_micro_average = list()#用于存储微平均

# 将特征和所对应的宏平均和微平均保存为excel文件
wb = Workbook()  # 创建工作簿
sheet = wb.active  # 获取到当前book获得的sheet页
row = 1  # 从第几行开始保存数据
col = 1  # 从第几列开始保存数据
for i in fs_num_list:
    print("=====卡方检验=====")
    sk =SelectKBest(chi2, k=i).fit(train_counts_indexs, train_target)#是用train_counts_indexs，还是train_tfidf_counts_indexs来进行特征选择。可打印print(index_train)来看出差别，用train_tfidf选择的特征主要集中在后面

    print("=====选择训练样本的tfidf值=====")
    index_train=sk.get_support(indices=True)
    sk_train_tfidf=train_tfidf_counts_indexs[:,index_train]#根据索引选择对应列的tfidf

    print("=====选择测试样本的tfidf值=====")
    index_test=sk.get_support(indices=True)
    sk_test_tfidf=test_tfidf_counts_indexs[:,index_test]
    print("=====使用朴素贝叶斯训练分类器=====")
    clf = MultinomialNB().fit(sk_train_tfidf,train_target)#alpha=0.01的结果为CHI2NB_20news_20180807，alpha=0.009的结果为

    print("=====测试集分类准确率=====")
    predicted = clf.predict(sk_test_tfidf)
    accuracy=np.mean(predicted == test_target)

    print("=====几种计算测试集分类准确率的方法=====")
    print("朴素贝叶斯-测试集分类准确率:"+str(accuracy))#或者将加号+改为逗号，就不会出现数据类型不同的问题。
    print(metrics.accuracy_score(test_target, predicted))#
    num_correct = np.sum(predicted == test_target)#注意这里用的sum



    print("=====朴素贝叶斯-各类别的精确度，召回率，F值的分类报告=====")
    print(metrics.classification_report(test_target,predicted,target_names = patent_target_names))

    print("=====根据分类报告计算精确度，召回率，F值的宏平均(macro average)=====")
    print("精确度的宏平均：",metrics.precision_score(test_target,predicted, average='macro'))
    print("召回率的宏平均：",metrics.recall_score(test_target,predicted, average='macro'))
    print("F值的宏平均：",metrics.f1_score(test_target,predicted, average='macro'))

    print("=====根据分类报告计算精确度，召回率，F值的微平均(micro average)=====")
    print("精确度的微平均：",metrics.precision_score(test_target,predicted, average='micro'))
    print("召回率的微平均：",metrics.recall_score(test_target,predicted, average='micro'))
    print("F值的微平均：",metrics.f1_score(test_target,predicted, average='micro'))

    print("=====将宏平均和微平均用列表存储=====")
    f1_macro_average.append(metrics.f1_score(test_target, predicted, average='macro'))
    f1_micro_average.append(metrics.f1_score(test_target, predicted, average='micro'))

    sheet.cell(row, col, i)  # col=0表示从第0列保存特征i
    sheet.cell(row, col + 1, metrics.f1_score(test_target, predicted, average='macro'))  # col+1=1表示从第1列保存特征i所对应的宏平均f1_macro_average
    sheet.cell(row, col + 2, metrics.f1_score(test_target, predicted, average='micro'))  # col+1=1表示从第1列保存特征i所对应的微平均f1_micro_average
    row += 1  # 列数保持不变，保存后行数增加1
wb.save('111.xlsx')  # 保存文件，该语句放在这里表示（没有放在缩进放在for循环中），所有特征i遍历完后才保存，共保存一次。
print("=====分别打印宏平均和微平均列表=====")
print(f1_macro_average)
print(f1_micro_average)

'''--------------（分类器选择SVM）用for循环从卡方检验中取不同的k个特征，即选出最优的k个特征用SVM进行文本分类----------------'''

fs_num_list1 = list(range(100, 1100, 100))#在特征范围10-100内，间隔10取特征，得到10-90的特征数。Python3中range() 函数返回的是一个可迭代对象（类型是对象），而不是列表类型， 所以打印的时候不会打印列表。因此要加上list
fs_num_list2 = list(range(1000, 4300, 500))#取所有特征
fs_num_list3 = [4200]#取所有特征
fs_num_list = fs_num_list1 +fs_num_list2+fs_num_list3 #合并特征列表_list3#合并特征列表


f1_macro_average = list()#用于存储宏平均
f1_micro_average = list()#用于存储微平均

# 将特征和所对应的宏平均和微平均保存为excel文件
wb = Workbook()  # 创建工作簿
sheet = wb.active  # 获取到当前book获得的sheet页
row = 1  # 从第几行开始保存数据
col = 1  # 从第几列开始保存数据
for i in fs_num_list:
    print("=====卡方检验=====")
    sk =SelectKBest(chi2, k=i).fit(train_counts_indexs, train_target)#是用train_counts_indexs，还是train_tfidf_counts_indexs来进行特征选择。可打印print(index_train)来看出差别，用train_tfidf选择的特征主要集中在后面

    print("=====选择训练样本的tfidf值=====")
    index_train=sk.get_support(indices=True)
    sk_train_tfidf=train_tfidf_counts_indexs[:,index_train]#根据索引选择对应列的tfidf

    print("=====选择测试样本的tfidf值=====")
    index_test=sk.get_support(indices=True)
    sk_test_tfidf=test_tfidf_counts_indexs[:,index_test]
#print(sk_test_tfidf.shape)
#print(sk_test_tfidf)

    print("当前正在对", i, "个特征用SVM分类器进行文本分类")  # 打印出来以晓得计算进度

    print("=====使用SVM训练分类器=====")
    clf = svm.SVC(kernel='linear').fit(sk_train_tfidf,train_target)#alpha=0.01的结果为CHI2NB_20news_20180807，alpha=0.009的结果为


    print("=====测试集分类准确率=====")
    predicted = clf.predict(sk_test_tfidf)
    accuracy=np.mean(predicted == test_target)

    print("=====几种计算测试集分类准确率的方法=====")
    print("SVM-测试集分类准确率:"+str(accuracy))#或者将加号+改为逗号，就不会出现数据类型不同的问题。
    print(metrics.accuracy_score(test_target, predicted))
    num_correct = np.sum(predicted == test_target)#注意这里用的sum


    print("=====SVM-各类别的精确度，召回率，F值的分类报告=====")
    print(metrics.classification_report(test_target,predicted,target_names = patent_target_names))

    print("=====根据分类报告计算精确度，召回率，F值的宏平均(macro average)=====")
    print("精确度的宏平均：",metrics.precision_score(test_target,predicted, average='macro'))
    print("召回率的宏平均：",metrics.recall_score(test_target,predicted, average='macro'))
    print("F值的宏平均：",metrics.f1_score(test_target,predicted, average='macro'))

    print("=====根据分类报告计算精确度，召回率，F值的微平均(micro average)=====")
    print("精确度的微平均：",metrics.precision_score(test_target,predicted, average='micro'))
    print("召回率的微平均：",metrics.recall_score(test_target,predicted, average='micro'))
    print("F值的微平均：",metrics.f1_score(test_target,predicted, average='micro'))

    print("=====将宏平均和微平均用列表存储=====")
    f1_macro_average.append(metrics.f1_score(test_target, predicted, average='macro'))
    f1_micro_average.append(metrics.f1_score(test_target, predicted, average='micro'))

    sheet.cell(row, col, i)  # col=0表示从第0列保存特征i
    sheet.cell(row, col + 1, metrics.f1_score(test_target, predicted, average='macro'))  # col+1=1表示从第1列保存特征i所对应的宏平均f1_macro_average
    sheet.cell(row, col + 2, metrics.f1_score(test_target, predicted, average='micro'))  # col+1=1表示从第1列保存特征i所对应的微平均f1_micro_average
    row += 1  # 列数保持不变，保存后行数增加1
wb.save('SVM.xlsx')  # 保存文件，该语句放在这里表示（没有放在缩进放在for循环中），所有特征i遍历完后才保存，共保存一次。
print("=====分别打印宏平均和微平均列表=====")
print(f1_macro_average)
print(f1_micro_average)

